# Nome do arquivo: extrator_corretoras2.py (versão com screenshot para debug)

from playwright.async_api import async_playwright
import time
import json
import httpx
from datetime import datetime
import os
import pandas as pd
import warnings
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import asyncio
from dotenv import load_dotenv
import logging
import random

# Configuração do logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filename='execucoes.log',
    filemode='a'
)
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
console_handler.setFormatter(formatter)
logging.getLogger().addHandler(console_handler)

load_dotenv()

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# === CONFIGURAÇÕES ===
USUARIO = os.getenv("ICATU_USUARIO")
SENHA = os.getenv("ICATU_SENHA")

if not USUARIO or not SENHA:
    logging.error("ERRO: As variáveis de ambiente ICATU_USUARIO e ICATU_SENHA devem ser definidas.")
    exit()

PASTA_DOWNLOAD = "downloads"
NOME_CORRETORA_MAE = "OUTLIER CORRETORA LTDA"
CNPJ_CORRETORA_MAE = "48.978.010/0001-22" # CNPJ da segunda opção
CUSTOM_HEADER_NAME = 'customHeader'
CONCURRENCY_LIMIT = 20

# === FUNÇÕES HELPERS E CLASSE DA API (sem alterações) ===
def to_utc_date(string_date):
    if not string_date: return None
    try:
        dt = datetime.fromisoformat(string_date.replace('Z', '+00:00'))
        return dt.strftime('%Y-%m-%d')
    except:
        return string_date
def join_array(arr, delimiter, prop):
    if arr and isinstance(arr, list):
        return delimiter.join([str(item.get(prop, '')) for item in arr])
    return ''
def to_formatted_line_of_business(value):
    if value == 'PREV': return 'Previdência'
    if value == 'VIDA': return 'Vida'
    return value
def to_product_status(customer_product):
    value = ''
    line = customer_product.get('linhaNegocio')
    if line == 'PREV': value = customer_product.get('situacaoCertificado')
    elif line == 'VIDA': value = customer_product.get('situacaoTitulo')
    if value == 'A': return 'Ativo'
    if value == 'C': return 'Cancelado'
    return value
class IcatuAPIClient:
    def __init__(self, token):
        self.token = token
        self.base_url = "https://portalcorretor.icatuseguros.com.br/casadocorretorgateway/api"
        self.headers = {
            'Authorization': token, 'Content-Type': 'application/json', CUSTOM_HEADER_NAME: ''
        }
        self.semaphore = asyncio.Semaphore(CONCURRENCY_LIMIT)
    async def _make_request(self, method, url, **kwargs):
        async with self.semaphore:
            async with httpx.AsyncClient(timeout=45) as client:
                try:
                    response = await client.request(method, url, headers=self.headers, **kwargs)
                    response.raise_for_status()
                    return response.json() if response.text else None
                except httpx.HTTPStatusError as e:
                    logging.error(f"Erro HTTP {e.response.status_code} em {method} {url}: {e.response.text}")
                    if e.response.status_code == 401: logging.warning("Token pode ter expirado.")
                    return None
                except (httpx.RequestError, json.JSONDecodeError, Exception) as e:
                    logging.error(f"Erro em {method} {url}: {e}")
                    await asyncio.sleep(2)
                    return None
    def _parse_cliente_unico(self, item, details):
        return {
            'id_cliente': details.get('codigoBaseAgrupada'), 'nome': details.get('nome'),
            'documento': f"{item.get('documento', {}).get('tipo')}: {item.get('documento', {}).get('numeroFormatado')}",
            'titular_cpf': details.get('titularCPF'), 'sexo': details.get('sexo'),
            'data_nascimento': details.get('dataNascimentoFormatada'), 'estado_civil': details.get('estadoCivilFormatado'),
            'tipo_documento': (details.get('identidade') or [{}])[0].get('tipoDocumento'),
            'numero_documento': (details.get('identidade') or [{}])[0].get('documento'),
            'orgao_expedidor': (details.get('identidade') or [{}])[0].get('orgaoExpedidor'),
            'renda_patrimonio': details.get('rendaResumidaFormatada'), 'profissao': details.get('profissao'),
            'telefone': join_array(details.get('telefone'), ';', 'numeroTelefone'),
            'email': (details.get('emails') or [{}])[0].get('email'),
            'endereco': (details.get('endereco') or [{}])[0].get('descricaoEndereco'),
            'numero': (details.get('endereco') or [{}])[0].get('numero'),
            'complemento': (details.get('endereco') or [{}])[0].get('complemento'),
            'bairro': (details.get('endereco') or [{}])[0].get('bairro'),
            'cidade': (details.get('endereco') or [{}])[0].get('municipio'),
            'uf': (details.get('endereco') or [{}])[0].get('uf'),
            'cep': (details.get('endereco') or [{}])[0].get('cepFormatado'),
        }
    def _parse_produto_prev(self, product, id_cliente):
        prod = {
            'id_cliente': id_cliente,
            'linha_negocio': to_formatted_line_of_business(product.get('linhaNegocio')),
            'tipo_produto': product.get('nomeProduto'),
            'numero_proposta': product.get('proposta'),
            'numero_certificado': product.get('certificado'),
            'valor_contribuicao': product.get('valorPagamento'),
            'situacao_produto': to_product_status(product),
            'numero_processo_susep': product.get('numeroProcessoSusep'),
            'dia_vencimento': product.get('diaVencimento'),
            'ultimo_pagamento': to_utc_date(product.get('dataUltimoPagamento')),
            'proximo_pagamento': to_utc_date(product.get('dataProximoPagamento')),
            'quantidade_parcelas_pagas': product.get('quantidadeParcelasPagas'),
            'quantidade_parcelas_pendentes': product.get('quantidadeParcelasPendentes'),
            'periodicidade_pagamentos': product.get('periodicidadePagamento'),
            'forma_pagamento': product.get('formaPagamento'),
        }
        acumulacao = product.get('prev', {}).get('acumulacao')
        if acumulacao:
            prod.update({
                'nome_fundo': acumulacao.get('fundo'),
                'cnpj_fundo': acumulacao.get('cnpjFundo'),
                'regime_tributario': acumulacao.get('regimeTribCertAcumulacao'),
                'indexador_plano': acumulacao.get('indexadorCertificadoAcumulacao'),
            })
        return prod
    def _parse_produto_vida(self, product, benefit, id_cliente):
        return {
            'id_cliente': id_cliente,
            'linha_negocio': to_formatted_line_of_business(product.get('linhaNegocio')),
            'tipo_produto': product.get('nomeProduto'),
            'numero_proposta': product.get('proposta'),
            'numero_certificado': product.get('certificado'),
            'situacao_produto': to_product_status(product),
            'nome_cobertura': benefit.get('nomeBeneficio'),
            'capital_segurado': benefit.get('capitalBeneficioSegurado'),
            'periodo_pagamento_cobertura': benefit.get('prazoPagamento'),
            'dia_vencimento': product.get('diaVencimento'),
            'ultimo_pagamento': to_utc_date(product.get('dataUltimoPagamento')),
            'proximo_pagamento': to_utc_date(product.get('dataProximoPagamento')),
            'quantidade_parcelas_pagas': product.get('quantidadeParcelasPagas'),
            'quantidade_parcelas_pendentes': product.get('quantidadeParcelasPendentes'),
            'periodicidade_pagamentos': product.get('periodicidadePagamento'),
        }
    async def get_customers(self, original_post_data, update_status_func=None):
        logging.info("Iniciando download de Clientes...")
        customers_list, page_count = [], 1
        while True:
            if update_status_func: update_status_func(f"Buscando página de clientes {page_count}...")
            post_data = json.loads(original_post_data); post_data['Pagina'] = page_count
            response_data = await self._make_request('POST', f"{self.base_url}/RelacionamentoCliente/Tombamento/clientes", json=post_data)
            if not response_data or not response_data.get('clientes'): break
            customers_list.extend(response_data['clientes'])
            page_count += 1
        logging.info(f"Total de {len(customers_list)} clientes encontrados...")
        async def get_customer_details(customer):
            details_url = f"{self.base_url}/RelacionamentoCliente/Tombamento/clientes/{customer['codigoBaseAgrupada']}"
            products_url = f"{self.base_url}/RelacionamentoCliente/Tombamento/clientes/{customer['codigoBaseAgrupada']}/produtos?documento={customer['cpfCnpj']}"
            details_data, products_data = await asyncio.gather(self._make_request('GET', details_url), self._make_request('GET', products_url))
            return customer, details_data, products_data
        results = await asyncio.gather(*[get_customer_details(c) for c in customers_list])
        clientes_unicos, produtos_prev, produtos_vida = {}, [], []
        for i, (item, details_res, products_res) in enumerate(results):
            if update_status_func: update_status_func(f"Processando cliente {i+1}/{len(results)}...")
            if not details_res or not products_res: continue
            details = details_res.get('detalhesCliente', {}).get('clientes', [{}])[0]
            id_cliente = details.get('codigoBaseAgrupada')
            if not id_cliente: continue
            if id_cliente not in clientes_unicos: clientes_unicos[id_cliente] = self._parse_cliente_unico(item, details)
            products = products_res.get('produtosCliente', {}).get('listarProdutos', [])
            for product in products:
                if product.get('linhaNegocio') == 'PREV': produtos_prev.append(self._parse_produto_prev(product, id_cliente))
                elif product.get('linhaNegocio') == 'VIDA':
                    for benefit in product.get('vida', {}).get('beneficios', []): produtos_vida.append(self._parse_produto_vida(product, benefit, id_cliente))
        return [{'name': 'Clientes', 'data': list(clientes_unicos.values())}, {'name': 'Produtos Previdencia', 'data': produtos_prev}, {'name': 'Produtos Vida', 'data': produtos_vida}]
    def _parse_pending_data(self, item):
        return {
            'linha_negocio': item.get("linhaNegocio"), 'produto': item.get("nomeProdutoComercial"), 'numero_proposta': item.get("numeroProposta"),
            'numero_certificado': item.get("numeroCertificado"), 'nome_cliente': item.get("nomeCliente"), 'cpf_cliente': item.get("cpfCnpjCliente"),
            'status_pagamento': item.get("statusPagamento"), 'vencimento_original': item.get("diaVencimentoOriginal"), 'vencimento_atual': item.get("diaVencimentoAtual"),
            'competencia': item.get("competencia"), 'forma_pagamento': item.get("formaCobranca"), 'contribuicao': item.get("valorParcela"),
            'dias_em_atraso': item.get("diasDeAtraso"), 'email_cliente': item.get("email"), 'telefone1': item.get("telefone1"), 'telefone2': item.get("telefone2"),
        }
    async def get_pending_payments(self, original_post_data, update_status_func=None):
        logging.info("Iniciando download de pagamentos pendentes...")
        final_data_list, page_count = [], 0
        while page_count < 1000:
            post_data = json.loads(original_post_data); post_data.update({'paginaAtual': page_count, 'tamanhoPagina': 100})
            if update_status_func: update_status_func(f"Buscando página de pendentes {page_count}...")
            response_data = await self._make_request('POST', f"{self.base_url}/Relatorio/pendentes/tabela/v2", json=post_data)
            client_list = response_data.get('pendentes', []) if response_data else []
            if not client_list: break
            final_data_list.extend([self._parse_pending_data(item) for item in client_list])
            page_count += 1
        logging.info(f"Total de {len(final_data_list)} registros pendentes encontrados.")
        return [{'name': 'Pagamentos Pendentes', 'data': final_data_list}]
    def _parse_proposal_status(self, item, installment):
        return {
            'nome': item.get('nomeProponente'), 'cpf': item.get('cpfProponente'), 'produto': item.get('nomeProduto'), 'linha_negocio': item.get('linhaNegocio'),
            'proposta': item.get('numeroProposta'), 'criada_em': item.get('dataProtocolo'), 'status_proposta': item.get('statusFase'), 'data': item.get('dataStatus'),
            'forma_pagamento': item.get('formaPagamento'), 'valor': installment.get('valor'), 'vencimento': installment.get('agendamentoDebito'),
            'competencia': installment.get('competencia'), 'status_pagamento': item.get('statusPagamento'), 'motivo_pendencia': item.get('motivoPendencia'),
        }
    async def get_proposal_status(self, original_post_data, update_status_func=None):
        logging.info("Iniciando download de Status de Propostas...")
        proposal_list, page_count = [], 1
        while True:
            if update_status_func: update_status_func(f"Buscando página de propostas {page_count}...")
            post_data = json.loads(original_post_data); post_data['Pagina'] = page_count
            response_data = await self._make_request('POST', f"{self.base_url}/relatorio/consulta/status/v2", json=post_data)
            if not response_data or not response_data.get('listaPropostas'): break
            proposal_list.extend(response_data['listaPropostas'])
            page_count += 1
        logging.info(f"Total de {len(proposal_list)} propostas encontradas...")
        async def get_proposal_details(proposal):
            url = f"{self.base_url}/Clientes/{proposal['cpfProponente']}/primeira-parcela/{proposal['numeroProposta']}/0"
            return proposal, await self._make_request('GET', url)
        results = await asyncio.gather(*[get_proposal_details(p) for p in proposal_list])
        final_data = [self._parse_proposal_status(item, details['resultado']) for item, details in results if details and details.get('resultado')]
        return [{'name': 'Status Propostas', 'data': final_data}]

def export_to_excel(filename, all_sheets_data):
    wb = Workbook()
    if 'Sheet' in wb.sheetnames: wb.remove(wb['Sheet'])
    for sheet_info in all_sheets_data:
        if not sheet_info['data']:
            logging.warning(f"Aviso: A aba '{sheet_info['name']}' está vazia, pulando...")
            continue
        try:
            ws = wb.create_sheet(title=sheet_info['name'])
            df = pd.DataFrame(sheet_info['data'])
            for r in dataframe_to_rows(df, index=False, header=True): ws.append(r)
            logging.info(f"Aba '{sheet_info['name']}' criada com sucesso.")
        except Exception as e: logging.error(f"Erro ao criar aba '{sheet_info['name']}': {e}")
    if not wb.sheetnames:
        logging.warning("Aviso: Nenhum dado foi retornado. O arquivo Excel não será gerado.")
        return
    try:
        wb.save(filename)
        logging.info(f"Arquivo Excel salvo: {filename}")
    except Exception as e: logging.error(f"Erro ao salvar arquivo Excel: {e}")

async def capture_post_data(page, target_url_part):
    request_log = []
    def handle_request(request):
        if target_url_part in request.url and request.method == 'POST' and request.post_data:
            if request.post_data not in request_log:
                request_log.append(request.post_data)

    page.on('request', handle_request)
    
    logging.info("Tentando acionar a busca de dados para capturar a comunicação...")
    try:
        await page.click('button:has-text("Buscar")', timeout=6000)
        await page.wait_for_load_state("networkidle", timeout=40000)
    except Exception:
        logging.warning("Botão 'Buscar' não encontrado, tentando recarregar a página como alternativa...")
        try:
            await page.reload(wait_until="networkidle")
        except Exception as e:
            logging.error(f"Falha ao recarregar a página: {e}")
    
    await asyncio.sleep(3) 
            
    page.remove_listener('request', handle_request)

    if request_log:
        logging.info("postData capturado com sucesso!")
        return request_log[0]
    
    logging.warning(f"Não foi possível capturar o postData para {target_url_part}")
    return None

async def aguardar_elemento_com_retry(page, selector, timeout=40000, max_attempts=4):
    for attempt in range(max_attempts):
        try:
            await page.wait_for_selector(selector, timeout=timeout)
            return True
        except Exception as e:
            logging.warning(f"Tentativa {attempt + 1} falhou para aguardar '{selector}': {e}")
            if attempt < max_attempts - 1: await asyncio.sleep(2)
    logging.error(f"ERRO: Elemento '{selector}' não encontrado após {max_attempts} tentativas.")
    return False

async def clicar_elemento_com_retry(page, selector, timeout=40000, max_attempts=4):
    for attempt in range(max_attempts):
        try:
            await page.wait_for_selector(selector, state='visible', timeout=timeout)
            await page.click(selector, timeout=timeout)
            logging.info(f"Clicou em: {selector}")
            return True
        except Exception as e:
            logging.warning(f"Tentativa {attempt + 1} de clique falhou para '{selector}': {e}")
            if attempt < max_attempts - 1: await asyncio.sleep(2)
    logging.error(f"ERRO: Não foi possível clicar em '{selector}' após {max_attempts} tentativas.")
    return False

# ##########################################################################
# ## FUNÇÃO MODIFICADA COM SCREENSHOT ##
# ##########################################################################
async def selecionar_corretora_por_cnpj(page, cnpj_desejado):
    """
    Seleciona uma corretora específica baseada no CNPJ, usando o seletor
    correto derivado do console.
    """
    # 1. Clica no elemento que abre a lista de corretoras
    seletor_dropdown = 'div.dsi_header-selected-item:has-text("Selecione")'
    logging.info(f"Abrindo a lista de corretoras clicando em: {seletor_dropdown}")
    if not await clicar_elemento_com_retry(page, seletor_dropdown):
        logging.error("Não foi possível abrir a lista de corretoras.")
        await page.screenshot(path="debug_falha_abrir_dropdown.png", full_page=True)
        return False

    # 2. Pausa para aguardar a animação do menu e o carregamento dos itens
    logging.info("Aguardando 2 segundos para a animação do menu...")
    await asyncio.sleep(2)

    # 3. TIRA A FOTO DA TELA PARA DEBUG
    screenshot_path = "debug_selecao_corretora.png"
    await page.screenshot(path=screenshot_path, full_page=True)
    logging.info(f"!!! Screenshot para debug salvo em: {screenshot_path} !!!")

    # 4. Cria um seletor específico baseado no HTML fornecido
    seletor_corretora = f'div.dsi_header-select-item-wrapper:has-text("{cnpj_desejado}")'
    logging.info(f"Tentando clicar na corretora com o seletor preciso: {seletor_corretora}")

    # 5. Tenta clicar diretamente na opção correta
    if await clicar_elemento_com_retry(page, seletor_corretora):
        logging.info(f"Corretora com CNPJ {cnpj_desejado} selecionada com sucesso.")
        await asyncio.sleep(1)  # Pausa para a UI atualizar após o clique
        return True
    else:
        logging.error(f"Não foi possível encontrar ou clicar na corretora com CNPJ {cnpj_desejado} usando o seletor preciso.")
        return False

async def processar_corretora(browser, corretora_info, corretora_index, total_corretoras):
    corretora_nome = corretora_info["nome"]
    logging.info(f"\n{'='*80}\nPROCESSANDO CORRETORA {corretora_index}/{total_corretoras}: {corretora_nome}\n{'='*80}")
    
    context = {"token": None, "token_captured": asyncio.Event()}

    async def intercept_token_response(response):
        if "/api/usuarios/corretoras" in response.url and "/contextualizar" in response.url:
            try:
                if response.ok:
                    json_body = await response.json()
                    token = json_body.get("resultado", {}).get("token")
                    if token:
                        context["token"] = f"Bearer {token}"
                        logging.info(f"\nTOKEN CAPTURADO PARA {corretora_nome}: ...{context['token'][-10:]}\n")
                        context["token_captured"].set()
                else:
                    response_text = await response.text()
                    logging.error(f"Erro na API de token para {corretora_nome}. Status: {response.status}. Resposta: {response_text}")

            except json.JSONDecodeError:
                response_text = await response.text()
                logging.error(f"Erro ao decodificar JSON da resposta de token para {corretora_nome}. Resposta recebida: {response_text}")
            except Exception as e:
                logging.error(f"Erro inesperado ao processar resposta para token: {e}")
                
    page = await browser.new_page()
    page.on("response", intercept_token_response)
    try:
        logging.info("Fase 1: Realizando login e selecionando corretora...")
        logging.info("Navegando para a página de login...")
        await page.goto("https://portalcorretor.icatuseguros.com.br/casadocorretor/login", wait_until="domcontentloaded", timeout=70000)
        
        try:
            await page.click('button#onetrust-accept-btn-handler', timeout=30000)
            logging.info("Banner de cookies aceito.")
        except Exception:
            logging.info("Banner de cookies não foi encontrado no tempo limite, continuando...")

        await aguardar_elemento_com_retry(page, 'input[placeholder="Usuário"]')
        logging.info("Preenchendo usuário e senha...")
        await page.fill('input[placeholder="Usuário"]', USUARIO)
        await page.fill('input[placeholder="Senha"]', SENHA)
        await clicar_elemento_com_retry(page, 'button.dsi-button-primary')

        logging.info("Aguardando o carregamento do portal após o login...")
        if not await aguardar_elemento_com_retry(page, 'div.dsi_header-selected-item:has-text("Selecione")', timeout=65000):
             logging.error("A página do portal não carregou o seletor de corretora a tempo.")
             return False

        logging.info(f"Selecionando corretora mãe com CNPJ: {CNPJ_CORRETORA_MAE}...")
        if not await selecionar_corretora_por_cnpj(page, CNPJ_CORRETORA_MAE):
            logging.error(f"Falha ao selecionar corretora mãe com CNPJ {CNPJ_CORRETORA_MAE}")
            return False
        
        await asyncio.sleep(2)

        if not await clicar_elemento_com_retry(page, 'button.dsi-button-link:has-text("Selecionar corretor vinculado a plataforma")'): return False
        
        logging.info(f"Selecionando a corretora específica: {corretora_nome}...")
        if not await clicar_elemento_com_retry(page, 'div.dsi_header-selected-item:has-text("Selecione")'): return False
        if not await clicar_elemento_com_retry(page, f'text="{corretora_nome}"'):
            logging.critical(f"ERRO CRÍTICO: Não foi possível encontrar/clicar na corretora '{corretora_nome}' na lista.")
            return False
        
        if not await clicar_elemento_com_retry(page, 'button:has-text("Selecionar")'): return False
        
        logging.info(f"Seleção para '{corretora_nome}' concluída. Aguardando captura do token...")
        try:
            await asyncio.wait_for(context["token_captured"].wait(), timeout=40.0)
        except asyncio.TimeoutError:
            logging.error(f"ERRO: Tempo esgotado esperando pelo token da corretora '{corretora_nome}'.")
            await page.reload(wait_until="networkidle")
            await asyncio.sleep(5)

        if not context["token"]:
            logging.error(f"Token não foi capturado para {corretora_nome}. Encerrando esta corretora.")
            await page.screenshot(path=f'debug_screenshot_{corretora_nome.replace(" ", "_")}_no_token.png')
            return False
            
        logging.info("Token obtido, iniciando extração de dados.")
        await page.wait_for_load_state("networkidle", timeout=60000)

        all_sheets_data = []
        api_client = IcatuAPIClient(context["token"])
        secoes = [
            ("Clientes", "/meus-clientes", '/api/RelacionamentoCliente/Tombamento/clientes', api_client.get_customers),
            ("Pagamentos Pendentes", "/meus-clientes/pendentes-beta", '/api/Relatorio/pendentes/tabela/v2', api_client.get_pending_payments),
            ("Status de Propostas", "/venda/status-proposta", '/api/relatorio/consulta/status/v2', api_client.get_proposal_status)
        ]
        logging.info("\nFase 2: Iniciando extração dos dados...")
        for nome, url_path, api_part, extract_func in secoes:
            logging.info(f"\n--- Extraindo {nome} para {corretora_nome} ---")
            await page.goto(f"https://portalcorretor.icatuseguros.com.br/casadocorretor{url_path}", wait_until="networkidle")
            post_data = await capture_post_data(page, api_part)
            if post_data:
                sheets = await extract_func(post_data, lambda s: logging.info(f"  Status: {s}"))
                all_sheets_data.extend(sheets)
            else:
                logging.warning(f"Não foi possível obter o postData para {nome}. Pulando seção.")
        
        logging.info("\nFase 3: Salvando arquivos locais...")
        if not os.path.exists(PASTA_DOWNLOAD): os.makedirs(PASTA_DOWNLOAD, exist_ok=True)
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        nome_base = f"Extracao_{corretora_nome.replace(' ', '_').replace('/', '-')}_{timestamp}"
        path_excel = os.path.join(PASTA_DOWNLOAD, f"{nome_base}.xlsx")
        export_to_excel(path_excel, all_sheets_data)
        path_json = os.path.join(PASTA_DOWNLOAD, f"{nome_base}_backup.json")
        try:
            with open(path_json, 'w', encoding='utf-8') as f:
                json.dump(all_sheets_data, f, ensure_ascii=False, indent=2, default=str)
            logging.info(f"Backup JSON salvo: {path_json}")
        except Exception as e: logging.error(f"Erro ao salvar backup JSON: {e}")
        
        logging.info(f"\nExtração concluída para {corretora_nome}!")
        for sheet in all_sheets_data: logging.info(f"  - {sheet['name']}: {len(sheet['data'])} registros")
        return True
    except Exception as e:
        logging.exception(f"ERRO CRÍTICO ao processar {corretora_nome}: {e}")
        await page.screenshot(path=f'debug_screenshot_{corretora_nome.replace(" ", "_")}.png')
        return False
    finally:
        if not page.is_closed():
            await page.close()
        logging.info(f"Sessão encerrada para {corretora_nome}")

async def main():
    start_time = time.time()
    
    arquivo_corretoras = 'corretoras_para_rerodar.xlsx'
    try:
        logging.info(f"Lendo a lista de corretoras do arquivo '{arquivo_corretoras}'...")
        df = pd.read_excel(arquivo_corretoras)
        df['nome'] = df['nome'].astype(str)
        
        corretoras_df = df[~df['nome'].str.contains(NOME_CORRETORA_MAE, case=False, na=False)]
        
        CORRETORAS = corretoras_df.to_dict('records')
        
        if len(df) > len(CORRETORAS):
            logging.warning(f"A corretora mãe '{NOME_CORRETORA_MAE}' foi encontrada no Excel e removida da lista de processamento.")

        logging.info(f"{len(CORRETORAS)} corretoras filhas carregadas para processamento.")

    except FileNotFoundError:
        logging.error(f"ERRO CRÍTICO: O arquivo '{arquivo_corretoras}' não foi encontrado.")
        logging.error("Por favor, crie o arquivo Excel com as colunas 'nome' e 'cnpj' e adicione as corretoras.")
        return
    except Exception as e:
        logging.error(f"ERRO CRÍTICO ao ler o arquivo Excel: {e}")
        return
    
    if not CORRETORAS:
        logging.warning("Nenhuma corretora filha para processar na lista. Encerrando.")
        return

    logging.info("INICIANDO PROCESSO DE EXTRAÇÃO DE DADOS...")
    if not os.path.exists(PASTA_DOWNLOAD):
        os.makedirs(PASTA_DOWNLOAD)
        logging.info(f"Pasta de downloads criada em: {os.path.abspath(PASTA_DOWNLOAD)}")
    
    sucessos = 0
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        for i, corretora in enumerate(CORRETORAS, 1):
            if await processar_corretora(browser, corretora, i, len(CORRETORAS)):
                sucessos += 1
            
            if i < len(CORRETORAS):
                delay = random.uniform(5, 15)
                logging.info(f"\nPausa de {delay:.2f} segundos antes da próxima corretora...")
                await asyncio.sleep(delay)

        await browser.close()
    
    end_time = time.time()
    logging.info(f"\n{'='*80}")
    logging.info(f"EXTRAÇÃO CONCLUÍDA PARA {sucessos}/{len(CORRETORAS)} CORRETORAS.")
    logging.info(f"TEMPO TOTAL: {end_time - start_time:.2f} SEGUNDOS.")
    logging.info(f"Os arquivos .xlsx e .json estão prontos na pasta '{PASTA_DOWNLOAD}'.")
    logging.info(f"{'='*80}")

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logging.warning("\nProcesso interrompido pelo usuário.")